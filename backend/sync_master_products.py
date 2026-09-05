import os
import json
from sqlalchemy import text
from app.core.database import SessionLocal, engine
from app.models.product import Product
from app.models.brand import Brand

def sync_products():
    json_path = os.path.join(os.path.dirname(__file__), "master_products_data.json")
    excel_path = r"C:\Icons\Desktop\feb to agu 31.xlsx"

    excel_items = []

    if os.path.exists(json_path):
        print(f"Loading master catalog from JSON: {json_path}...")
        with open(json_path, "r", encoding="utf-8") as f:
            excel_items = json.load(f)
    elif os.path.exists(excel_path):
        print(f"Loading Excel: {excel_path} (Sheet3)...")
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['Sheet3']

        def infer_spec(size_str, name_str, code_str, cat_str):
            size_u = str(size_str or '').upper().strip()
            name_u = str(name_str or '').upper().strip()
            code_u = str(code_str or '').upper().strip()
            cat_u = str(cat_str or '').upper().strip()

            category = cat_u
            if not category:
                if code_u == 'B': category = 'BRANDY'
                elif code_u == 'R': category = 'RUM'
                elif code_u == 'W': category = 'WHISKY'
                elif code_u == 'V': category = 'VODKA'
                elif code_u == 'WI': category = 'WINE'
                elif code_u == 'GIN': category = 'GIN'
                elif code_u == 'IFL': category = 'IFL'
                elif 'BEER' in size_u or 'BEER' in name_u: category = 'BEER'
                elif 'BRANDY' in name_u or 'BDY' in name_u: category = 'BRANDY'
                elif 'RUM' in name_u: category = 'RUM'
                elif 'WHISKY' in name_u or 'WISKY' in name_u or 'WHY' in name_u: category = 'WHISKY'
                elif 'VODKA' in name_u: category = 'VODKA'
                elif 'WINE' in name_u: category = 'WINE'
                elif 'GIN' in name_u: category = 'GIN'
                else: category = 'SPIRITS'

            volume_ml = 750
            pack_size = 12

            if 'BEER' in size_u or 'BEER' in category or 'BEER' in name_u:
                if '500' in name_u: volume_ml = 500; pack_size = 24
                elif '330' in name_u: volume_ml = 330; pack_size = 24
                else: volume_ml = 650; pack_size = 12
            elif '1LIT' in size_u or '1000' in name_u or '1L' in name_u:
                volume_ml = 1000
                pack_size = 12
            elif 'HALF' in size_u or '375' in name_u:
                volume_ml = 375
                pack_size = 24
            elif 'QUATER' in size_u or 'QUARTER' in size_u or '180' in name_u or name_u.endswith(' Q'):
                volume_ml = 180
                pack_size = 48
            elif 'FULL' in size_u:
                volume_ml = 750
                pack_size = 12

            clean_name = str(name_str).strip()
            clean_u = clean_name.upper()

            if size_u == 'QUATER' or size_u == 'QUARTER':
                if not any(k in clean_u for k in ['QUARTER', 'QUATER', ' Q', ' QTR', ' 180ML']):
                    clean_name = f"{clean_name} QUARTER"
            elif size_u == 'HALF':
                if not any(k in clean_u for k in ['HALF', ' H', ' 375ML']):
                    clean_name = f"{clean_name} HALF"
            elif size_u == 'FULL':
                if not any(k in clean_u for k in ['FULL', ' F', ' 750ML']):
                    clean_name = f"{clean_name} FULL"
            elif size_u == 'FULL 1LIT':
                if not any(k in clean_u for k in ['1LIT', '1000ML', ' 1L']):
                    clean_name = f"{clean_name} 1L"

            return clean_name, category, volume_ml, pack_size

        for r in range(3, ws.max_row + 1):
            id_val = ws.cell(r, 1).value
            size_val = ws.cell(r, 2).value
            name_val = ws.cell(r, 3).value
            code_val = ws.cell(r, 4).value
            cat_val = ws.cell(r, 5).value
            basic_val = ws.cell(r, 6).value
            mrp_val = ws.cell(r, 7).value
            sales_val = ws.cell(r, 8).value
            
            if id_val == '*' or basic_val is None or name_val is None:
                continue

            clean_name, cat, vol, pack = infer_spec(size_val, name_val, code_val, cat_val)
            
            excel_items.append({
                'name': clean_name,
                'category': cat,
                'volume_ml': vol,
                'pack_size': pack,
                'basic_rate': round(float(basic_val), 2),
                'mrp': int(round(float(mrp_val))),
                'selling_price': int(round(float(sales_val)))
            })

    if not excel_items:
        print("No product catalog items found to sync.")
        return

    print(f"Loaded {len(excel_items)} items for catalog sync.")

    db = SessionLocal()
    
    # Ensure default brands exist for categories
    brands_by_cat = {}
    for cat in set(item['category'] for item in excel_items):
        brand = db.query(Brand).filter(Brand.category == cat).first()
        if not brand:
            brand = Brand(name=f"Standard {cat.title()} Brand", category=cat, is_active=True, is_deleted=False)
            db.add(brand)
            db.commit()
            db.refresh(brand)
        brands_by_cat[cat] = brand.id

    existing_products = db.query(Product).all()

    # Index existing products by (name_lower, volume_ml) and name_lower
    by_key = {(p.name.strip().lower(), p.volume_ml): p for p in existing_products}
    by_name = {p.name.strip().lower(): p for p in existing_products}

    processed_ids = set()
    updated_count = 0
    created_count = 0

    for item in excel_items:
        clean_name = item['name'].strip()
        lower_name = clean_name.lower()
        vol = item['volume_ml']
        brand_id = brands_by_cat[item['category']]

        p = by_key.get((lower_name, vol)) or by_name.get(lower_name)

        if p and p.id not in processed_ids:
            p.name = clean_name
            p.category = item['category']
            p.volume_ml = vol
            p.pack_size = item['pack_size']
            p.basic_rate = item['basic_rate']
            p.mrp = item['mrp']
            p.selling_price = item['selling_price']
            p.is_active = True
            p.is_deleted = False
            p.brand_id = brand_id
            processed_ids.add(p.id)
            updated_count += 1
        else:
            new_p = Product(
                name=clean_name,
                category=item['category'],
                volume_ml=vol,
                unit='bottle',
                pack_size=item['pack_size'],
                basic_rate=item['basic_rate'],
                mrp=item['mrp'],
                selling_price=item['selling_price'],
                brand_id=brand_id,
                is_active=True,
                is_deleted=False
            )
            db.add(new_p)
            db.flush()
            processed_ids.add(new_p.id)
            created_count += 1

    # Soft delete all products not in master catalog
    deactivated_count = 0
    for p in db.query(Product).all():
        if p.id not in processed_ids:
            p.is_active = False
            p.is_deleted = True
            deactivated_count += 1

    db.commit()
    db.close()

    print(f"[OK] SYNC COMPLETED SUCCESSFULLY!")
    print(f"Updated existing products: {updated_count}")
    print(f"Created new products: {created_count}")
    print(f"Deactivated/Soft-deleted products: {deactivated_count}")
    print(f"Active products remaining: {len(processed_ids)}")

if __name__ == '__main__':
    sync_products()
