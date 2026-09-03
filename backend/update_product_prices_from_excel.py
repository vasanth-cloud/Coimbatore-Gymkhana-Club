import openpyxl
from sqlalchemy import text
from app.core.database import SessionLocal, engine
from app.models.product import Product

excel_path = r"C:\Icons\Desktop\CGC SEPTEMBER 26.xlsx"
print(f"Loading Excel file: {excel_path} (Sheet: '01.09')...")

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['01.09']

db = SessionLocal()
db_products = db.query(Product).filter(Product.is_deleted == False).all()

print(f"Found {len(db_products)} products in PostgreSQL database.")

updated_count = 0
unmatched = []

for r in range(4, ws.max_row + 1):
    vol_val = ws.cell(r, 1).value
    cat_val = ws.cell(r, 2).value
    desc_val = ws.cell(r, 3).value
    mrp_val = ws.cell(r, 4).value
    sales_val = ws.cell(r, 5).value

    if not desc_val or 'TOTAL' in str(desc_val).upper():
        continue

    try:
        mrp_num = int(float(str(mrp_val))) if mrp_val is not None else 0
        sales_num = int(float(str(sales_val))) if sales_val is not None else 0
        vol_num = int(float(str(vol_val))) if vol_val is not None else None
    except Exception:
        continue

    if sales_num == 0 and mrp_num == 0:
        continue

    desc_clean = str(desc_val).strip().lower()

    # Search for matching product in PostgreSQL
    matched_p = None

    # Strategy 1: Exact / substring match with volume
    for p in db_products:
        p_name_clean = p.name.lower()
        if vol_num and p.volume_ml == vol_num:
            if desc_clean in p_name_clean or p_name_clean.startswith(desc_clean):
                matched_p = p
                break

    # Strategy 2: Match without volume constraint
    if not matched_p:
        for p in db_products:
            p_name_clean = p.name.lower()
            if desc_clean in p_name_clean or p_name_clean.startswith(desc_clean):
                matched_p = p
                break

    if matched_p:
        matched_p.mrp = mrp_num
        matched_p.selling_price = sales_num
        matched_p.basic_rate = mrp_num  # Set Basic Purchase Rate to MRP rate
        
        # Calculate pack size
        if matched_p.volume_ml <= 180:
            matched_p.pack_size = 48
        elif matched_p.volume_ml == 375:
            matched_p.pack_size = 24
        else:
            matched_p.pack_size = 12

        updated_count += 1
    else:
        unmatched.append(f"{vol_num}ml {cat_val} {desc_val} (MRP: ₹{mrp_num}, Sales: ₹{sales_num})")

db.commit()

print(f"✅ UPDATED PRICES FOR {updated_count} PRODUCTS IN POSTGRESQL!")
print(f"Unmatched items count: {len(unmatched)}")
if unmatched:
    print("Sample unmatched items:", unmatched[:10])
